from app.main import bp
import json
import csv
import json
import io
from flask import render_template, request, jsonify, send_file
from app.models.attachment import Attachment
from app.extensions import db
from io import BytesIO
import xlrd, mmap
from openpyxl import load_workbook


last_inserted_id = 0

# This route handles the homepage
@bp.route('/')
def home():
    return render_template('index.html')

# This route processes uploaded CSV, XLS, or XLSX files and converts them to JSON format
# and saves the JSON data in a PostgreSQL database
@bp.route('/process_csv', methods=['POST'])
def to_json():
    # This function will be triggered when a POST request is made to the '/process_csv' endpoint
    # It expects the file to be uploaded as part of the request
    # The function access the uploaded file using request.files and perform the necessary conversion operations
    
    global last_inserted_id

    # Vérifier si un fichier a été soumis
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier n\'a été soumis.'}), 400
    
    file = request.files['file']

     # Check if the file size is small and equal to 1 GB
    if file.content_length <= 1 * 1024 * 1024 * 1024:
        # Perform file handling or processing here for the specific file size
        # Vérifier si le fichier est un fichier CSV
        if file.filename.endswith('.csv'):
            try:
                jsonString = csv_to_json(file)
                filename = file.filename
            except Exception as e:
                return jsonify({'error': 'Fichier non valide ou Erreur lors de la lecture du fichier CSV.'}), 400
            
            try:
                #Add file to database
                attachment = Attachment(path = filename, content=jsonString)
                db.session.add(attachment)
                db.session.commit()

                # Retrieve the last inserted ID
                last_inserted_id = attachment.id
                
            except Exception as e:
                return jsonify({'error': 'Fichier non valide ou Erreur lors de la création du fichier JSON de sortie.'}), 500
            
            return render_template('result.html', attachment_id=last_inserted_id)
        
        elif file.filename.endswith('.xls'):        
            try:
                jsonString = excel_to_json(file)

                filename = file.filename
                
            except Exception as e:
                return jsonify({'error': 'Fichier non valide ou Erreur lors de la lecture du fichier EXCEL.'}), 400
            
            try:

                #Add file to database
                attachment = Attachment(path = filename, content=jsonString)
                db.session.add(attachment)
                db.session.commit()

                # Retrieve the last inserted ID
                last_inserted_id = attachment.id
                
            except Exception as e:
                return jsonify({'error': 'Fichier non valide ou Erreur lors de la création du fichier JSON de sortie.'}), 500
            
            return render_template('result.html', attachment_id=last_inserted_id)
        
        elif file.filename.endswith('.xlsx'): 
            try:
                jsonString = excel_xlsx_to_json(file)

                filename = file.filename
                
            except Exception as e:
                return jsonify({'error': 'Fichier non valide ou Erreur lors de la lecture du fichier EXCEL.'}), 400
            
            try:

                #Add file to database
                attachment = Attachment(path = filename, content=jsonString)
                db.session.add(attachment)
                db.session.commit()

                # Retrieve the last inserted ID
                last_inserted_id = attachment.id
                
            except Exception as e:
                return jsonify({'error': 'Fichier non valide ou Erreur lors de la création du fichier JSON de sortie.'}), 500

            #To rende the 'result.html' template with the appropriate context
            return render_template('result.html', attachment_id=last_inserted_id)
        else:
            return jsonify({'error': 'Le fichier soumis n\'est pas un fichier CSV ou EXCEL.'}), 400
    else:
        return jsonify({'error': 'Fichier dépassant 1 GB.'}), 400
    

# The function download() is defined as a route handler for 
# the endpoint /download. It can handle both GET and POST requests.
@bp.route('/download', methods=['GET', 'POST'])
def download():

    # Retrieves an Attachment object from the database based on the provided attachment_id
    attachment  = Attachment.query.filter_by(id=last_inserted_id).first()

    # The 'send_file' function to send the file as a response to the client
    if attachment:    
        return send_file(BytesIO(attachment.content.encode()), 
                         mimetype='application/json', 
                         download_name="yourfile.json",
                         as_attachment=True)
    else:
        return jsonify({'error': 'Aucun fichier JSON à télécharger.'}), 400


# The function api_json_generated() is defined as a route handler for 
# the endpoint /api_json_generated/<attachment_id>/json/lojl. It can handle both GET and POST requests.
@bp.route('/leckomba-jude/api_json_generated/<attachment_id>/json/lojl', methods=['GET', 'POST'])
def api_json_generated(attachment_id):
    
    # Retrieves an Attachment object from the database based on the provided attachment_id
    attachment  = Attachment.query.filter_by(id=attachment_id).first()

    # The 'send_file' function to send the file as a response to the client
    if attachment:    
        return send_file(BytesIO(attachment.content.encode()), 
                         mimetype='application/json', 
                         download_name="yourfile.json",
                         as_attachment=True)
    else:
        return jsonify({'error': 'Aucun fichier JSON à télécharger.'}), 400


def excel_to_json(file):
    # read .xls for excel format
    # Open the workbook using xlrd module to read the contents of the file
    workbook = xlrd.open_workbook(file_contents=file.read())
    
    # Access the first sheet of the workbook
    sheet = workbook.sheet_by_index(0)

    data = []

    # Extract the header row values as keys for the JSON data
    keys = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
    
    # Iterate over the rows starting from the second row
    for row in range(1, sheet.nrows):
        # Create a dictionary for each row, with keys from the header row and values from the current row
        row_data = {keys[col]: sheet.cell_value(row, col) for col in range(sheet.ncols)}
        data.append(row_data)

    # Convert the data list to JSON format using the json module
    # Indent the JSON data for better readability
    json_data = json.dumps(data, indent=4)

    return json_data


def csv_to_json(file):
    # Read .csv file
    # Wrap the file object with TextIOWrapper to handle encoding
    with io.TextIOWrapper(file, encoding='utf-8') as csvfile:
        
        # Use csv.Sniffer to automatically detect the dialect of the CSV file
        dialect = csv.Sniffer().sniff(csvfile.read(1024))
                
        # Reset the file position to the beginning
        csvfile.seek(0)

        # Create a CSV reader object using DictReader to read the CSV file as dictionaries
        reader = csv.DictReader(csvfile, dialect=dialect)

        # Create a CSV reader object using DictReader to read the CSV file as dictionaries
        data = [dict(row) for row in reader]
    
    # Convert the list of dictionaries to JSON format using the json module
    # Indent the JSON data for better readability
    json_data = json.dumps(data, indent=4)

    return json_data


def excel_xlsx_to_json(file):
    # Read .xlsx file
    # Load the workbook using the openpyxl module
    workbook = load_workbook(filename=file)
    # Access the active sheet of the workbook
    sheet = workbook.active

    data = []
    
    # Extract the header row values as the keys for the JSON data
    header = [cell.value for cell in sheet[1]]
    
    # Iterate over the rows starting from the second row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Create a dictionary for each row, with keys from the header row and values from the current row
        row_data = dict(zip(header, row))
        data.append(row_data)

    # Convert the data list to JSON format using the json module
    # Indent the JSON data for better readability
    return json.dumps(data, indent=4)